#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified order analytics engine used by both the Flask app and the skill.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from copy import deepcopy
from datetime import date, datetime
from difflib import SequenceMatcher
from io import BytesIO, TextIOWrapper
import csv
import json
import re
from xml.etree import ElementTree as ET
from typing import Dict, Iterable, List, Tuple, Union
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.exceptions import InvalidFileException

REQUIRED_SKU_FIELDS = (
    "seller_sku",
    "order_substatus",
    "cancel_type",
    "shipped_time",
    "created_time",
)
REGION_FIELD = "region"
METRIC_KEYS = (
    "total",
    "completed",
    "delivered",
    "refund",
    "cancel_before",
    "cancel_after",
    "in_transit",
)
UNKNOWN_LIMIT = 20
HEADER_SCAN_LIMIT = 8
BUCKET_LABELS = {
    "completed": "已完成",
    "delivered": "已送达",
    "refund": "退款",
    "cancel_before": "发货前取消",
    "cancel_after": "发货后取消",
    "in_transit": "仍在途",
    "unknown_status": "未知状态",
}

DEFAULT_MAPPING_PRESETS = {
    "cn_en": {
        "columns": {
            "seller_sku": ["seller sku", "seller_sku", "sku", "商家sku", "seller sku id", "variation"],
            "order_substatus": ["order substatus", "order sub status", "订单子状态", "子状态"],
            "cancel_type": [
                "cancelation/return type",
                "cancellation/return type",
                "cancelation return type",
                "cancellation return type",
                "取消/退货类型",
                "退货类型",
            ],
            "shipped_time": ["shipped time", "shipping time", "发货时间", "出货时间"],
            "created_time": ["created time", "creation time", "订单创建时间", "创建时间"],
            REGION_FIELD: [
                "province",
                "state",
                "province/state",
                "state/province",
                "region",
                "province name",
                "地区",
                "省份",
                "州",
                "区域",
            ],
        },
        "status_aliases": {
            "completed": ["已完成", "completed"],
            "delivered": ["已送达", "delivered"],
            "cancelled": ["已取消", "canceled", "cancelled", "cancel"],
            "in_transit": [
                "运输中",
                "in transit",
                "待揽收",
                "待发货",
                "未支付",
                "awaiting pickup",
                "awaiting collection",
                "to ship",
                "pending shipment",
                "pending payment",
                "unpaid",
            ],
            "refund": ["return/refund", "refund", "return", "退款", "退货"],
        },
        "return_aliases": ["return/refund", "refund", "return", "退款", "退货"],
        "cancel_aliases": ["cancel", "canceled", "cancelled", "已取消"],
        "region_aliases": [
            "province",
            "state",
            "province/state",
            "state/province",
            "region",
            "地区",
            "省份",
            "州",
            "区域",
        ],
    }
}


class MappingDetectionError(ValueError):
    """Raised when the engine cannot identify required columns."""

    def __init__(self, message: str, diagnostics: dict, suggestion: dict):
        super().__init__(message)
        self.diagnostics = diagnostics
        self.suggestion = suggestion


def _norm(text) -> str:
    return str(text).strip().lower() if text is not None else ""


def _norm_header(text) -> str:
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = value.replace("_", " ").replace("-", " ")
    value = value.replace("/", " / ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _tokenize(value: str) -> set:
    return {token for token in re.split(r"[^0-9a-zA-Z\u4e00-\u9fff]+", value) if token}


def _to_date(val) -> Union[date, None]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(val).date()
        except Exception:
            return None
    if isinstance(val, str):
        txt = val.strip()
        txt = txt.replace("年", "-").replace("月", "-").replace("日", "")
        patterns = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %I:%M %p",
            "%d/%m/%Y %I:%M:%S %p",
            "%d/%m/%Y %I:%M %p",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in patterns:
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                continue
        match = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", txt)
        if match:
            try:
                return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            except ValueError:
                return None
    return None


def _date_in_range(d: Union[datetime, date, str, int, float, None], start: date, end: date) -> bool:
    parsed = _to_date(d)
    if parsed is None:
        return False
    return start <= parsed <= end


def _deep_merge(base: dict, override: dict) -> dict:
    merged = deepcopy(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = value
    return merged


def load_mapping_config(preset: str = "cn_en", mapping_path: str | None = None, mapping_override: dict | None = None) -> dict:
    if preset not in DEFAULT_MAPPING_PRESETS:
        raise KeyError(f"未知映射预设: {preset}")

    config = deepcopy(DEFAULT_MAPPING_PRESETS[preset])

    override = {}
    if mapping_path:
        with open(mapping_path, "r", encoding="utf-8") as file_obj:
            override = json.load(file_obj)
    if mapping_override:
        override = _deep_merge(override, mapping_override)

    config = _deep_merge(config, override)
    if "region_aliases" in config:
        region_aliases = list(dict.fromkeys(config["region_aliases"]))
        column_aliases = config["columns"].setdefault(REGION_FIELD, [])
        config["columns"][REGION_FIELD] = list(dict.fromkeys(region_aliases + column_aliases))
    return config


def _xlsx_raw_bytes(data: Union[str, bytes, BytesIO]) -> bytes:
    if isinstance(data, str):
        with open(data, "rb") as file_obj:
            return file_obj.read()
    if isinstance(data, bytes):
        return data
    data.seek(0)
    return data.read()


def _repair_sparse_xlsx_rows(raw: bytes) -> Tuple[List[List], str] | None:
    try:
        with ZipFile(BytesIO(raw)) as archive:
            worksheet_names = [
                name for name in archive.namelist()
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            ]
            if not worksheet_names:
                return None

            shared_strings = []
            if "xl/sharedStrings.xml" in archive.namelist():
                shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for item in shared_root.findall("m:si", ns):
                    shared_strings.append("".join(t.text or "" for t in item.findall(".//m:t", ns)))

            ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            root = ET.fromstring(archive.read(worksheet_names[0]))
            cells = []
            max_row = 0
            max_col = 0
            for cell in root.findall(".//m:sheetData/m:row/m:c", ns):
                ref = cell.get("r")
                if not ref:
                    continue
                row_idx, col_idx = coordinate_to_tuple(ref)
                value_node = cell.find("m:v", ns)
                inline_node = cell.find("m:is", ns)
                value = None
                if cell.get("t") == "s" and value_node is not None and value_node.text is not None:
                    shared_idx = int(value_node.text)
                    value = shared_strings[shared_idx] if shared_idx < len(shared_strings) else None
                elif cell.get("t") == "inlineStr" and inline_node is not None:
                    value = "".join(t.text or "" for t in inline_node.findall(".//m:t", ns))
                elif value_node is not None:
                    value = value_node.text
                cells.append((row_idx, col_idx, value))
                max_row = max(max_row, row_idx)
                max_col = max(max_col, col_idx)

            if max_row <= 1 or max_col <= 1:
                return None
            rows = [[None for _ in range(max_col)] for _ in range(max_row)]
            for row_idx, col_idx, value in cells:
                rows[row_idx - 1][col_idx - 1] = value
            return rows, "OrderSKUList"
    except Exception:
        return None


def _load_rows_from_xlsx(data: Union[str, bytes, BytesIO]) -> Tuple[List[List], str]:
    raw = _xlsx_raw_bytes(data)
    workbook = load_workbook(BytesIO(raw), data_only=True, read_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    title = worksheet.title
    workbook.close()
    if len(rows) <= 1 and (not rows or len(rows[0]) <= 1):
        repaired = _repair_sparse_xlsx_rows(raw)
        if repaired is not None:
            return repaired
    return rows, title


def _load_rows_from_csv(data: BytesIO) -> Tuple[List[List], str]:
    data.seek(0)
    wrapper = TextIOWrapper(data, encoding="utf-8-sig")
    try:
        rows = [row for row in csv.reader(wrapper)]
    finally:
        wrapper.detach()
    return rows, "CSV"


def _load_rows(source: Union[str, bytes, BytesIO]) -> Tuple[List[List], str]:
    if isinstance(source, str):
        with open(source, "rb") as file_obj:
            raw = file_obj.read()
        return _load_rows(BytesIO(raw))
    if isinstance(source, bytes):
        return _load_rows(BytesIO(source))

    source.seek(0)
    data = BytesIO(source.read())
    try:
        rows, title = _load_rows_from_xlsx(data)
    except (InvalidFileException, BadZipFile):
        rows, title = _load_rows_from_csv(data)
    return rows, title


def _build_header_lookup(headers: List[str]) -> Dict[str, int]:
    return {
        _norm_header(header): idx
        for idx, header in enumerate(headers)
        if header is not None and str(header).strip()
    }


def _match_columns(headers: List[str], mapping_config: dict, required_fields: Iterable[str]) -> Tuple[Dict[str, int], Dict[str, str]]:
    header_lookup = _build_header_lookup(headers)
    positions: Dict[str, int] = {}
    matched_columns: Dict[str, str] = {}
    for field in required_fields:
        aliases = mapping_config["columns"].get(field, [])
        for alias in aliases:
            normalized_alias = _norm_header(alias)
            if normalized_alias in header_lookup:
                positions[field] = header_lookup[normalized_alias]
                matched_columns[field] = str(headers[header_lookup[normalized_alias]])
                break
    return positions, matched_columns


def _optional_column(headers: List[str], aliases: Iterable[str]) -> int | None:
    header_lookup = _build_header_lookup(headers)
    for alias in aliases:
        idx = header_lookup.get(_norm_header(alias))
        if idx is not None:
            return idx
    return None


def _row_value(row: List, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _sku_value(row: List, seller_idx: int, fallback_idx: int | None = None):
    seller_sku = _row_value(row, seller_idx)
    if seller_sku is not None and str(seller_sku).strip() != "":
        return seller_sku
    return _row_value(row, fallback_idx)


def _score_header_row(row: List, mapping_config: dict, required_fields: Iterable[str]) -> Tuple[float, Dict[str, int], Dict[str, str]]:
    positions, matched = _match_columns(row, mapping_config, required_fields)
    score = float(len(positions))
    if "seller_sku" in positions:
        score += 0.5
    if "created_time" in positions:
        score += 0.5
    return score, positions, matched


def detect_header_row(rows: List[List], mapping_config: dict, required_fields: Iterable[str]) -> Tuple[int, Dict[str, int], Dict[str, str]]:
    best = (0.0, 0, {}, {})
    scan_rows = rows[:HEADER_SCAN_LIMIT]
    for row_index, row in enumerate(scan_rows, start=1):
        score, positions, matched = _score_header_row(row, mapping_config, required_fields)
        if score > best[0]:
            best = (score, row_index, positions, matched)
    return best[1] or 1, best[2], best[3]


def _field_candidates(headers: List[str], aliases: List[str], limit: int = 5) -> List[str]:
    candidates = []
    normalized_aliases = [_norm_header(alias) for alias in aliases]
    alias_tokens = [_tokenize(alias) for alias in normalized_aliases]
    for header in headers:
        header_text = "" if header is None else str(header)
        normalized_header = _norm_header(header_text)
        header_tokens = _tokenize(normalized_header)
        best_score = 0.0
        for alias, tokens in zip(normalized_aliases, alias_tokens):
            ratio = SequenceMatcher(None, normalized_header, alias).ratio()
            overlap = 0.0
            if tokens and header_tokens:
                overlap = len(tokens & header_tokens) / len(tokens | header_tokens)
            best_score = max(best_score, ratio, overlap)
        candidates.append((best_score, header_text))
    candidates.sort(key=lambda item: (-item[0], item[1]))
    return [header for score, header in candidates if header][:limit]


def build_mapping_suggestion(headers: List[str], mapping_config: dict, missing_fields: Iterable[str]) -> dict:
    columns = {}
    diagnostics = {}
    for field in missing_fields:
        aliases = mapping_config["columns"].get(field, [])
        candidates = _field_candidates(headers, aliases)
        diagnostics[field] = candidates
        if candidates:
            columns[field] = [candidates[0]]
    return {
        "preset": "cn_en",
        "columns": columns,
        "status_aliases": mapping_config.get("status_aliases", {}),
        "return_aliases": mapping_config.get("return_aliases", []),
        "cancel_aliases": mapping_config.get("cancel_aliases", []),
        "region_aliases": mapping_config.get("columns", {}).get(REGION_FIELD, []),
        "candidate_columns": diagnostics,
    }


def _ensure_columns(headers: List[str], mapping_config: dict, positions: Dict[str, int], matched_columns: Dict[str, str], required_fields: Iterable[str], file_label: str, header_row: int) -> Tuple[Dict[str, int], Dict[str, str]]:
    missing_fields = [field for field in required_fields if field not in positions]
    if not missing_fields:
        return positions, matched_columns

    suggestion = build_mapping_suggestion(headers, mapping_config, missing_fields)
    diagnostics = {
        "file": file_label,
        "header_row": header_row,
        "missing_fields": missing_fields,
        "matched_columns": matched_columns,
        "candidate_columns": suggestion["candidate_columns"],
    }
    missing_names = ", ".join(missing_fields)
    raise MappingDetectionError(f"未识别到必要字段: {missing_names}", diagnostics, suggestion)


def classify_status(order_substatus, cancel_type, shipped_time, mapping_config: dict) -> str:
    substatus = _norm(order_substatus)
    cancel = _norm(cancel_type)
    shipped_empty = shipped_time is None or str(shipped_time).strip() == ""

    status_aliases = mapping_config["status_aliases"]
    return_aliases = {_norm(alias) for alias in mapping_config.get("return_aliases", [])}
    cancel_aliases = {_norm(alias) for alias in mapping_config.get("cancel_aliases", [])}

    completed_set = {_norm(alias) for alias in status_aliases.get("completed", [])}
    delivered_set = {_norm(alias) for alias in status_aliases.get("delivered", [])}
    canceled_set = {_norm(alias) for alias in status_aliases.get("cancelled", [])}
    in_transit_set = {_norm(alias) for alias in status_aliases.get("in_transit", [])}
    refund_set = {_norm(alias) for alias in status_aliases.get("refund", [])}

    if cancel in return_aliases or substatus in refund_set or "return" in substatus or "refund" in substatus:
        return "refund"
    if substatus in canceled_set or cancel in cancel_aliases:
        return "cancel_before" if shipped_empty else "cancel_after"
    if substatus in delivered_set:
        return "delivered"
    if substatus in completed_set and cancel == "":
        return "completed"
    if substatus in in_transit_set:
        return "in_transit"
    return "unknown_status"


def _empty_metrics() -> Dict[str, int]:
    return {key: 0 for key in METRIC_KEYS}


def _init_result_container() -> dict:
    return {
        "sku_stats": defaultdict(_empty_metrics),
        "region_stats": defaultdict(lambda: defaultdict(_empty_metrics)),
        "sku_totals": defaultdict(int),
        "diagnostics": {
            "files": [],
            "unknown_statuses": [],
            "matched_columns": {},
        },
        "summary": {
            "total_files": 0,
            "total_orders": 0,
            "sku_count": 0,
            "region_count": 0,
            "month_count": 0,
        },
    }


def _init_daily_container() -> dict:
    return {
        "sku_stats": defaultdict(_empty_metrics),
        "region_stats": defaultdict(lambda: defaultdict(_empty_metrics)),
        "sku_totals": defaultdict(int),
        "processed_orders": Counter(),
        "file_unknown": defaultdict(Counter),
        "overall_unknown": Counter(),
    }


def _merge_metrics(target: Dict[str, int], source: Dict[str, int]) -> None:
    for key in METRIC_KEYS:
        target[key] += int(source.get(key, 0))


def _metric_row(name: str, total: int, metrics: Dict[str, int]) -> dict:
    if total == 0:
        return {
            name: "",
            "total": 0,
            "sign_rate": 0.0,
            "completed_rate": 0.0,
            "delivered_rate": 0.0,
            "refund_rate": 0.0,
            "cancel_before_rate": 0.0,
            "cancel_after_rate": 0.0,
            "in_transit_rate": 0.0,
        }

    completed_rate = metrics.get("completed", 0) / total * 100
    delivered_rate = metrics.get("delivered", 0) / total * 100
    refund_rate = metrics.get("refund", 0) / total * 100
    cancel_before_rate = metrics.get("cancel_before", 0) / total * 100
    cancel_after_rate = metrics.get("cancel_after", 0) / total * 100
    in_transit_rate = metrics.get("in_transit", 0) / total * 100

    return {
        name: "",
        "total": total,
        "sign_rate": round(completed_rate + delivered_rate, 2),
        "completed_rate": round(completed_rate, 2),
        "delivered_rate": round(delivered_rate, 2),
        "refund_rate": round(refund_rate, 2),
        "cancel_before_rate": round(cancel_before_rate, 2),
        "cancel_after_rate": round(cancel_after_rate, 2),
        "in_transit_rate": round(in_transit_rate, 2),
    }


def build_sku_rows(sku_stats: Dict[str, Dict[str, int]]) -> List[dict]:
    rows = []
    for seller_sku, metrics in sorted(sku_stats.items(), key=lambda item: (-item[1]["total"], item[0])):
        total = metrics["total"]
        if total == 0:
            continue
        row = _metric_row("seller_sku", total, metrics)
        row["seller_sku"] = seller_sku
        rows.append(row)
    return rows


def build_region_rows(region_stats: Dict[str, Dict[str, Dict[str, int]]], sku_totals: Dict[str, int]) -> List[dict]:
    rows = []
    for seller_sku, region_map in sorted(region_stats.items(), key=lambda item: item[0]):
        total_sku = sku_totals.get(seller_sku, 0)
        for region, metrics in sorted(region_map.items(), key=lambda item: (-item[1]["total"], item[0])):
            total = metrics["total"]
            row = _metric_row("region", total, metrics)
            row["seller_sku"] = seller_sku
            row["region"] = region
            row["share_rate"] = round(total / total_sku * 100, 2) if total_sku else 0.0
            rows.append(row)
    return rows


def build_monthly_rows(monthly_stats: Dict[str, Dict[str, int]]) -> List[dict]:
    rows = []
    for month, metrics in sorted(monthly_stats.items(), key=lambda item: item[0]):
        total = metrics["total"]
        if total == 0:
            continue
        row = _metric_row("month", total, metrics)
        row["month"] = month
        rows.append(row)
    return rows


def build_monthly_sku_rows(monthly_sku_stats: Dict[str, Dict[str, Dict[str, int]]]) -> Dict[str, List[dict]]:
    return {
        month: build_sku_rows(sku_stats)
        for month, sku_stats in sorted(monthly_sku_stats.items(), key=lambda item: item[0])
    }


def build_daily_rows(daily_stats: Dict[str, Dict[str, int]], daily_sku_rows: Dict[str, List[dict]]) -> List[dict]:
    rows = []
    for day, metrics in sorted(daily_stats.items(), key=lambda item: item[0]):
        total = metrics["total"]
        if total == 0:
            continue
        row = _metric_row("date", total, metrics)
        row["date"] = day
        row["sku_count"] = len(daily_sku_rows.get(day, []))
        rows.append(row)
    return rows


def build_daily_sku_rows(daily_sku_stats: Dict[str, Dict[str, Dict[str, int]]]) -> Dict[str, List[dict]]:
    return {
        day: build_sku_rows(sku_stats)
        for day, sku_stats in sorted(daily_sku_stats.items(), key=lambda item: item[0])
    }


def _build_workbook(headers: List[str], data_rows: List[List], title: str) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    worksheet.append(headers)
    for row in data_rows:
        worksheet.append(row)
    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = 16
    worksheet.freeze_panes = "A2"
    return workbook


def build_sku_workbook(sku_rows: List[dict]) -> Workbook:
    headers = [
        "Seller SKU",
        "订单数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    data = [
        [
            row["seller_sku"],
            row["total"],
            row["sign_rate"],
            row["completed_rate"],
            row["delivered_rate"],
            row["refund_rate"],
            row["cancel_before_rate"],
            row["cancel_after_rate"],
            row["in_transit_rate"],
        ]
        for row in sku_rows
    ]
    return _build_workbook(headers, data, "订单指标")


def build_region_workbook(region_rows: List[dict]) -> Workbook:
    headers = [
        "Seller SKU",
        "Region",
        "订单数",
        "订单占比(%)",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    data = [
        [
            row["seller_sku"],
            row["region"],
            row["total"],
            row["share_rate"],
            row["sign_rate"],
            row["completed_rate"],
            row["delivered_rate"],
            row["refund_rate"],
            row["cancel_before_rate"],
            row["cancel_after_rate"],
            row["in_transit_rate"],
        ]
        for row in region_rows
    ]
    return _build_workbook(headers, data, "地区指标")


def build_monthly_workbook(monthly_rows: List[dict], monthly_sku_rows: Dict[str, List[dict]] | None = None) -> Workbook:
    headers = [
        "月份",
        "订单数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    data = [
        [
            row["month"],
            row["total"],
            row["sign_rate"],
            row["completed_rate"],
            row["delivered_rate"],
            row["refund_rate"],
            row["cancel_before_rate"],
            row["cancel_after_rate"],
            row["in_transit_rate"],
        ]
        for row in monthly_rows
    ]
    workbook = _build_workbook(headers, data, "月度总览")
    sku_headers = [
        "Seller SKU",
        "订单数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]

    for month, sku_rows in (monthly_sku_rows or {}).items():
        worksheet = workbook.create_sheet(title=month[:31])
        worksheet.append(sku_headers)
        for row in sku_rows:
            worksheet.append(
                [
                    row["seller_sku"],
                    row["total"],
                    row["sign_rate"],
                    row["completed_rate"],
                    row["delivered_rate"],
                    row["refund_rate"],
                    row["cancel_before_rate"],
                    row["cancel_after_rate"],
                    row["in_transit_rate"],
                ]
            )
        for column_index in range(1, len(sku_headers) + 1):
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = 16
        worksheet.freeze_panes = "A2"

    return workbook


def build_daily_workbook(daily_rows: List[dict], daily_sku_rows: Dict[str, List[dict]]) -> Workbook:
    overview_headers = [
        "日期",
        "订单数",
        "SKU数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    overview_data = [
        [
            row["date"],
            row["total"],
            row["sku_count"],
            row["sign_rate"],
            row["completed_rate"],
            row["delivered_rate"],
            row["refund_rate"],
            row["cancel_before_rate"],
            row["cancel_after_rate"],
            row["in_transit_rate"],
        ]
        for row in daily_rows
    ]
    workbook = _build_workbook(overview_headers, overview_data, "日度总览")

    detail = workbook.create_sheet(title="日度SKU明细")
    detail_headers = [
        "日期",
        "Seller SKU",
        "订单数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    detail.append(detail_headers)
    for day, sku_rows in daily_sku_rows.items():
        for row in sku_rows:
            detail.append(
                [
                    day,
                    row["seller_sku"],
                    row["total"],
                    row["sign_rate"],
                    row["completed_rate"],
                    row["delivered_rate"],
                    row["refund_rate"],
                    row["cancel_before_rate"],
                    row["cancel_after_rate"],
                    row["in_transit_rate"],
                ]
            )
    for column_index in range(1, len(detail_headers) + 1):
        detail.column_dimensions[get_column_letter(column_index)].width = 16
    detail.freeze_panes = "A2"

    notes = workbook.create_sheet(title="口径说明")
    notes.append(["字段", "说明"])
    notes.append(["签收率(%)", "已完成率 + 已送达率"])
    notes.append(["退款率(%)", "单独展示，不计入签收率"])
    notes.append(["日期", "按 Created Time 的本地日期聚合"])
    notes.append(["日度SKU明细", "每一天每个 Seller SKU 一行，用于透视表或折线趋势"])
    for column_index in range(1, 3):
        notes.column_dimensions[get_column_letter(column_index)].width = 32
    notes.freeze_panes = "A2"

    return workbook


def build_structured_workbook(structured_rows: List[dict]) -> Workbook:
    headers = [
        "文件",
        "Created Date",
        "月份",
        "Seller SKU",
        "Region",
        "状态分组",
        "状态分组代码",
        "是否计入签收",
        "未知 Order Substatus",
        "未知 Cancelation/Return Type",
    ]
    data = []
    signed_buckets = {"completed", "delivered"}
    for row in structured_rows:
        bucket = row.get("bucket", "")
        unknown_status = row.get("unknown_status") or {}
        created_date = row.get("created_date", "")
        data.append(
            [
                row.get("file", ""),
                created_date,
                created_date[:7],
                row.get("seller_sku", ""),
                row.get("region", ""),
                BUCKET_LABELS.get(bucket, bucket),
                bucket,
                "是" if bucket in signed_buckets else "否",
                unknown_status.get("order_substatus", ""),
                unknown_status.get("cancel_type", ""),
            ]
        )
    return _build_workbook(headers, data, "结构化订单明细")


def analyze_order_files(
    file_sources: Iterable[Union[str, bytes, BytesIO]],
    start_date: date,
    end_date: date,
    preset: str = "cn_en",
    mapping_path: str | None = None,
    mapping_override: dict | None = None,
    require_region: bool = False,
) -> dict:
    mapping_config = load_mapping_config(preset=preset, mapping_path=mapping_path, mapping_override=mapping_override)
    result = _init_result_container()
    result["diagnostics"]["preset"] = preset
    unknown_counter: Counter = Counter()
    required_fields = list(REQUIRED_SKU_FIELDS) + ([REGION_FIELD] if require_region else [])

    for index, source in enumerate(file_sources, start=1):
        file_label = getattr(source, "name", None) or f"file_{index}"
        rows, sheet_title = _load_rows(source)
        if not rows:
            continue

        header_row, positions, matched_columns = detect_header_row(rows, mapping_config, required_fields)
        headers = rows[header_row - 1]
        positions, matched_columns = _ensure_columns(
            headers=headers,
            mapping_config=mapping_config,
            positions=positions,
            matched_columns=matched_columns,
            required_fields=required_fields,
            file_label=file_label,
            header_row=header_row,
        )

        processed_rows = 0
        file_unknown = Counter()
        sku_fallback_idx = _optional_column(headers, ["Variation", "SKU Variation", "Product Variation"])
        for row in rows[header_row:]:
            seller_sku = _sku_value(row, positions["seller_sku"], sku_fallback_idx)
            created_time = row[positions["created_time"]] if positions["created_time"] < len(row) else None
            if seller_sku is None or str(seller_sku).strip() == "":
                continue
            if not _date_in_range(created_time, start_date, end_date):
                continue

            order_substatus = row[positions["order_substatus"]] if positions["order_substatus"] < len(row) else None
            cancel_type = row[positions["cancel_type"]] if positions["cancel_type"] < len(row) else None
            shipped_time = row[positions["shipped_time"]] if positions["shipped_time"] < len(row) else None
            region = ""
            if REGION_FIELD in positions and positions[REGION_FIELD] < len(row) and row[positions[REGION_FIELD]] is not None:
                region = str(row[positions[REGION_FIELD]]).strip()

            bucket = classify_status(order_substatus, cancel_type, shipped_time, mapping_config)
            seller_sku = str(seller_sku).strip()
            sku_metrics = result["sku_stats"][seller_sku]
            sku_metrics["total"] += 1
            result["sku_totals"][seller_sku] += 1
            processed_rows += 1

            if bucket == "unknown_status":
                unknown_key = (_norm(order_substatus), _norm(cancel_type))
                unknown_counter[unknown_key] += 1
                file_unknown[unknown_key] += 1
            else:
                sku_metrics[bucket] += 1

            if REGION_FIELD in positions:
                region_metrics = result["region_stats"][seller_sku][region]
                region_metrics["total"] += 1
                if bucket != "unknown_status":
                    region_metrics[bucket] += 1

        result["summary"]["total_orders"] += processed_rows
        result["diagnostics"]["files"].append(
            {
                "file": file_label,
                "sheet": sheet_title,
                "header_row": header_row,
                "matched_columns": matched_columns,
                "processed_orders": processed_rows,
                "unknown_statuses": [
                    {
                        "order_substatus": order_substatus,
                        "cancel_type": cancel_type,
                        "count": count,
                    }
                    for (order_substatus, cancel_type), count in file_unknown.most_common(UNKNOWN_LIMIT)
                ],
            }
        )

    sku_rows = build_sku_rows(result["sku_stats"])
    region_rows = build_region_rows(result["region_stats"], result["sku_totals"])
    result["sku_rows"] = sku_rows
    result["region_rows"] = region_rows
    result["summary"]["sku_count"] = len(sku_rows)
    result["summary"]["region_count"] = len(region_rows)
    result["summary"]["total_files"] = len(result["diagnostics"]["files"])
    result["diagnostics"]["unknown_statuses"] = [
        {
            "order_substatus": order_substatus,
            "cancel_type": cancel_type,
            "count": count,
        }
        for (order_substatus, cancel_type), count in unknown_counter.most_common(UNKNOWN_LIMIT)
    ]
    if result["diagnostics"]["files"]:
        result["diagnostics"]["matched_columns"] = result["diagnostics"]["files"][0]["matched_columns"]
    return result


def prepare_order_cache(
    file_sources: Iterable[Union[str, bytes, BytesIO]],
    preset: str = "cn_en",
    mapping_path: str | None = None,
    mapping_override: dict | None = None,
    require_region: bool = False,
) -> dict:
    mapping_config = load_mapping_config(preset=preset, mapping_path=mapping_path, mapping_override=mapping_override)
    required_fields = list(REQUIRED_SKU_FIELDS) + ([REGION_FIELD] if require_region else [])
    prepared = {
        "preset": preset,
        "date_basis": "Created Time",
        "start_date": None,
        "end_date": None,
        "files": [],
        "matched_columns": {},
        "normalized_rows": [],
        "daily": defaultdict(_init_daily_container),
    }

    for index, source in enumerate(file_sources, start=1):
        file_label = getattr(source, "name", None) or f"file_{index}"
        rows, sheet_title = _load_rows(source)
        if not rows:
            continue

        header_row, positions, matched_columns = detect_header_row(rows, mapping_config, required_fields)
        headers = rows[header_row - 1]
        positions, matched_columns = _ensure_columns(
            headers=headers,
            mapping_config=mapping_config,
            positions=positions,
            matched_columns=matched_columns,
            required_fields=required_fields,
            file_label=file_label,
            header_row=header_row,
        )

        if not prepared["matched_columns"]:
            prepared["matched_columns"] = matched_columns

        created_idx = positions["created_time"]
        seller_idx = positions["seller_sku"]
        sku_fallback_idx = _optional_column(headers, ["Variation", "SKU Variation", "Product Variation"])
        file_start = None
        file_end = None
        dated_rows = 0

        for row in rows[header_row:]:
            seller_sku = _sku_value(row, seller_idx, sku_fallback_idx)
            if seller_sku is None or str(seller_sku).strip() == "":
                continue

            created_time = row[created_idx] if created_idx < len(row) else None
            created_date = _to_date(created_time)
            if created_date is None:
                continue

            order_substatus = row[positions["order_substatus"]] if positions["order_substatus"] < len(row) else None
            cancel_type = row[positions["cancel_type"]] if positions["cancel_type"] < len(row) else None
            shipped_time = row[positions["shipped_time"]] if positions["shipped_time"] < len(row) else None
            region = ""
            if REGION_FIELD in positions and positions[REGION_FIELD] < len(row) and row[positions[REGION_FIELD]] is not None:
                region = str(row[positions[REGION_FIELD]]).strip()

            bucket = classify_status(order_substatus, cancel_type, shipped_time, mapping_config)
            seller_sku = str(seller_sku).strip()
            day_bucket = prepared["daily"][created_date]
            sku_metrics = day_bucket["sku_stats"][seller_sku]
            sku_metrics["total"] += 1
            day_bucket["sku_totals"][seller_sku] += 1
            day_bucket["processed_orders"][file_label] += 1
            dated_rows += 1

            normalized_row = {
                "file": file_label,
                "created_date": created_date.isoformat(),
                "seller_sku": seller_sku,
                "region": region,
                "bucket": bucket,
            }

            if bucket == "unknown_status":
                unknown_key = (_norm(order_substatus), _norm(cancel_type))
                day_bucket["overall_unknown"][unknown_key] += 1
                day_bucket["file_unknown"][file_label][unknown_key] += 1
                normalized_row["unknown_status"] = {
                    "order_substatus": unknown_key[0],
                    "cancel_type": unknown_key[1],
                }
            else:
                sku_metrics[bucket] += 1

            if REGION_FIELD in positions:
                region_metrics = day_bucket["region_stats"][seller_sku][region]
                region_metrics["total"] += 1
                if bucket != "unknown_status":
                    region_metrics[bucket] += 1

            prepared["normalized_rows"].append(normalized_row)

            if file_start is None or created_date < file_start:
                file_start = created_date
            if file_end is None or created_date > file_end:
                file_end = created_date

        if file_start is not None and (prepared["start_date"] is None or file_start < prepared["start_date"]):
            prepared["start_date"] = file_start
        if file_end is not None and (prepared["end_date"] is None or file_end > prepared["end_date"]):
            prepared["end_date"] = file_end

        prepared["files"].append(
            {
                "file": file_label,
                "sheet": sheet_title,
                "header_row": header_row,
                "matched_columns": matched_columns,
                "dated_rows": dated_rows,
                "start_date": file_start.isoformat() if file_start else None,
                "end_date": file_end.isoformat() if file_end else None,
            }
        )

    return prepared


def analyze_prepared_order_cache(prepared: dict, start_date: date, end_date: date) -> dict:
    result = _init_result_container()
    result["diagnostics"]["preset"] = prepared.get("preset", "cn_en")
    result["diagnostics"]["matched_columns"] = prepared.get("matched_columns", {})
    overall_unknown = Counter()
    file_processed: Counter = Counter()
    file_unknown: defaultdict[str, Counter] = defaultdict(Counter)
    monthly_stats: defaultdict[str, Dict[str, int]] = defaultdict(_empty_metrics)
    monthly_sku_stats: defaultdict[str, defaultdict[str, Dict[str, int]]] = defaultdict(lambda: defaultdict(_empty_metrics))
    daily_stats: defaultdict[str, Dict[str, int]] = defaultdict(_empty_metrics)
    daily_sku_stats: defaultdict[str, defaultdict[str, Dict[str, int]]] = defaultdict(lambda: defaultdict(_empty_metrics))
    structured_rows = []

    for current_date, day_bucket in prepared.get("daily", {}).items():
        if not (start_date <= current_date <= end_date):
            continue

        month_key = current_date.strftime("%Y-%m")
        day_key = current_date.isoformat()
        for seller_sku, metrics in day_bucket["sku_stats"].items():
            _merge_metrics(result["sku_stats"][seller_sku], metrics)
            _merge_metrics(monthly_stats[month_key], metrics)
            _merge_metrics(monthly_sku_stats[month_key][seller_sku], metrics)
            _merge_metrics(daily_stats[day_key], metrics)
            _merge_metrics(daily_sku_stats[day_key][seller_sku], metrics)
        for seller_sku, total in day_bucket["sku_totals"].items():
            result["sku_totals"][seller_sku] += total
        for seller_sku, region_map in day_bucket["region_stats"].items():
            for region, metrics in region_map.items():
                _merge_metrics(result["region_stats"][seller_sku][region], metrics)
        for file_label, processed in day_bucket["processed_orders"].items():
            file_processed[file_label] += processed
            result["summary"]["total_orders"] += processed
        for file_label, unknown_counter in day_bucket["file_unknown"].items():
            file_unknown[file_label].update(unknown_counter)
        overall_unknown.update(day_bucket["overall_unknown"])

    for row in prepared.get("normalized_rows", []):
        row_date = datetime.strptime(row["created_date"], "%Y-%m-%d").date()
        if start_date <= row_date <= end_date:
            structured_rows.append(row)

    for file_info in prepared.get("files", []):
        file_label = file_info["file"]
        result["diagnostics"]["files"].append(
            {
                "file": file_label,
                "sheet": file_info["sheet"],
                "header_row": file_info["header_row"],
                "matched_columns": file_info["matched_columns"],
                "processed_orders": file_processed.get(file_label, 0),
                "unknown_statuses": [
                    {
                        "order_substatus": order_substatus,
                        "cancel_type": cancel_type,
                        "count": count,
                    }
                    for (order_substatus, cancel_type), count in file_unknown[file_label].most_common(UNKNOWN_LIMIT)
                ],
            }
        )

    result["sku_rows"] = build_sku_rows(result["sku_stats"])
    result["region_rows"] = build_region_rows(result["region_stats"], result["sku_totals"])
    result["monthly_rows"] = build_monthly_rows(monthly_stats)
    result["monthly_sku_rows"] = build_monthly_sku_rows(monthly_sku_stats)
    result["daily_sku_rows"] = build_daily_sku_rows(daily_sku_stats)
    result["daily_rows"] = build_daily_rows(daily_stats, result["daily_sku_rows"])
    result["structured_rows"] = structured_rows
    result["summary"]["sku_count"] = len(result["sku_rows"])
    result["summary"]["region_count"] = len(result["region_rows"])
    result["summary"]["month_count"] = len(result["monthly_rows"])
    result["summary"]["total_files"] = len(prepared.get("files", []))
    result["diagnostics"]["unknown_statuses"] = [
        {
            "order_substatus": order_substatus,
            "cancel_type": cancel_type,
            "count": count,
        }
        for (order_substatus, cancel_type), count in overall_unknown.most_common(UNKNOWN_LIMIT)
    ]
    return result


def write_mapping_suggestion(path: str, suggestion: dict) -> None:
    with open(path, "w", encoding="utf-8") as file_obj:
        json.dump(suggestion, file_obj, ensure_ascii=False, indent=2)
